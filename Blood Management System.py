import os
import time
from datetime import datetime
from collections import deque

from colorama import init, Fore, Style
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from tabulate import tabulate
import msvcrt

# ---------------------------- Configuration ----------------------------
init(autoreset=True)

DB_FILENAME = "blood_management.xlsx"
LOW_STOCK_THRESHOLD = 5
MAX_BAR_LENGTH = 30
DASHBOARD_REFRESH_SECONDS = 1
HISTORY_LIMIT = 20
BLOOD_TYPES = ["A+", "A-", "B+", "B-", "AB+", "AB-", "O+", "O-"]

donation_trend = deque(maxlen=HISTORY_LIMIT)
usage_trend = deque(maxlen=HISTORY_LIMIT)

# ---------------------------- Banner ----------------------------
def print_banner():
    banner = r"""
  ____  _                 _     __  __                  _             
 | __ )| | ___   __ _  __| |   |  \/  | ___  _ __   ___| | _____ _ __ 
 |  _ \| |/ _ \ / _` |/ _` |   | |\/| |/ _ \| '_ \ / __| |/ / _ \ '__|
 | |_) | | (_) | (_| | (_| |   | |  | | (_) | | | | (__|   <  __/ |   
 |____/|_|\___/ \__,_|\__,_|   |_|  |_|\___/|_| |_|\___|_|\_\___|_|   
"""
    print(Fore.CYAN + Style.BRIGHT + banner)

# ---------------------------- Excel Utilities ----------------------------
def initialize_database():
    if os.path.exists(DB_FILENAME):
        return

    wb = Workbook()

    # Donors
    ws = wb.active
    ws.title = "Donors"
    ws.append(["Donor ID", "Name", "Contact", "Age", "Blood Type", "Blood Units Donated", "Date"])
    format_sheet(ws)

    # Patients
    ws = wb.create_sheet("Patients")
    ws.append(["Patient ID", "Name", "Contact", "Age", "Blood Type", "Blood Units Needed", "Date"])
    format_sheet(ws)

    # Blood Stock
    ws = wb.create_sheet("BloodStock")
    ws.append(["Blood Type", "Units"])
    for b in BLOOD_TYPES:
        ws.append([b, 0])
    format_sheet(ws)

    # History
    ws = wb.create_sheet("History")
    ws.append(["ID", "Type", "Name", "Blood Type", "Units", "Action", "Date"])
    format_sheet(ws)

    wb.save(DB_FILENAME)


def format_sheet(ws):
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border


def load_workbook_safe():
    return load_workbook(DB_FILENAME)


def generate_id(sheet_name):
    wb = load_workbook_safe()
    ws = wb[sheet_name]
    last_id = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and str(row[0]).isdigit():
            last_id = max(last_id, int(row[0]))

    return last_id + 1

# ---------------------------- Blood Stock & History ----------------------------
def update_blood_stock(blood_type, units, action="add"):
    wb = load_workbook_safe()
    ws = wb["BloodStock"]

    for row in ws.iter_rows(min_row=2):
        if row[0].value == blood_type:
            try:
                current_units = int(row[1].value)
            except:
                current_units = 0

            if action == "add":
                row[1].value = current_units + units
            elif action == "subtract":
                row[1].value = max(0, current_units - units)
            break

    wb.save(DB_FILENAME)


def log_history(entity_id, entity_type, name, blood_type, units, action):
    wb = load_workbook_safe()
    ws = wb["History"]

    ws.append([
        entity_id,
        entity_type,
        name,
        blood_type,
        units,
        action,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ])
    wb.save(DB_FILENAME)

# ---------------------------- Excel Dashboard ----------------------------
def update_excel_dashboard():
    wb = load_workbook(DB_FILENAME)

    if "Dashboard" in wb.sheetnames:
        del wb["Dashboard"]

    ws_dash = wb.create_sheet("Dashboard")

    # Title
    ws_dash.merge_cells("A1:E1")
    title = ws_dash["A1"]
    title.value = "BLOOD BANK DASHBOARD SUMMARY"
    title.font = Font(bold=True, size=16, color="FFFFFF")
    title.fill = PatternFill("solid", fgColor="4F81BD")
    title.alignment = Alignment(horizontal="center")

    # Stock Table Header
    ws_dash.append(["Blood Type", "Units Available"])

    ws_stock = wb["BloodStock"]

    # Add Data
    for row in ws_stock.iter_rows(min_row=2, values_only=True):
        ws_dash.append(list(row))

    # Style + Fix: Convert to int first
    for row in ws_dash.iter_rows(min_row=2, max_col=2):
        try:
            units = int(row[1].value)
        except:
            units = 0

        for cell in row:
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(bold=True)

        # Colors
        if units <= LOW_STOCK_THRESHOLD:
            row[1].fill = PatternFill("solid", fgColor="FF0000")  # Red
        elif units <= LOW_STOCK_THRESHOLD * 2:
            row[1].fill = PatternFill("solid", fgColor="FFFF00")  # Yellow

    # Bar Chart
    chart = BarChart()
    chart.title = "Blood Stock by Type"
    data = Reference(ws_dash, min_col=2, min_row=2, max_row=ws_stock.max_row + 1)
    cats = Reference(ws_dash, min_col=1, min_row=3, max_row=ws_stock.max_row + 1)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 15
    ws_dash.add_chart(chart, "D3")

    # Donation vs Usage Table
    ws_hist = wb["History"]
    donations = {}
    usage = {}

    for row in ws_hist.iter_rows(min_row=2, values_only=True):
        date = row[6][:10]
        units = int(row[4])

        if row[1] == "Donor":
            donations[date] = donations.get(date, 0) + units
        elif row[1] == "Patient":
            usage[date] = usage.get(date, 0) + units

    ws_dash.append([])
    ws_dash.append(["Date", "Units Donated", "Units Used"])

    start = ws_dash.max_row

    for d in sorted(set(list(donations.keys()) + list(usage.keys()))):
        ws_dash.append([d, donations.get(d, 0), usage.get(d, 0)])

    # Line Chart
    line = LineChart()
    line.title = "Daily Donations vs Usage"
    data = Reference(ws_dash, min_col=2, min_row=start, max_col=3, max_row=ws_dash.max_row)
    cats = Reference(ws_dash, min_col=1, min_row=start+1, max_row=ws_dash.max_row)
    line.add_data(data, titles_from_data=True)
    line.set_categories(cats)
    line.height = 10
    line.width = 20
    ws_dash.add_chart(line, "E20")

    ws_dash.freeze_panes = "A2"

    wb.save(DB_FILENAME)

# ---------------------------- Console Dashboard ----------------------------
def display_live_dashboard():
    wb = load_workbook_safe()
    ws = wb["BloodStock"]

    stock = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            stock[row[0]] = int(row[1])
        except:
            stock[row[0]] = 0

    os.system('cls')
    print_banner()
    print(Fore.CYAN + "=== Live Blood Stock Dashboard ===\n")

    for b, u in stock.items():
        bar_len = int(u * MAX_BAR_LENGTH / 50)
        bar = "█" * bar_len

        if u <= LOW_STOCK_THRESHOLD:
            color = Fore.RED
        elif u <= LOW_STOCK_THRESHOLD * 2:
            color = Fore.YELLOW
        else:
            color = Fore.GREEN

        print(f"{b:>3}: {color}{bar} {u} units")

    print("\nTrend Graphs:")
    print(Fore.MAGENTA + "Donations: " + "".join("█" * (x // 2) for x in donation_trend))
    print(Fore.YELLOW + "Usage:     " + "".join("█" * (x // 2) for x in usage_trend))

    print(Fore.CYAN + "\nD=Donor | P=Patient | H=History | S=Search | U=Update | X=Delete | Q=Quit\n")

# ---------------------------- Animations ----------------------------
def animate_message(msg, color=Fore.GREEN):
    for ch in msg:
        print(color + ch, end="", flush=True)
        time.sleep(0.02)
    print()

# ---------------------------- CRUD ----------------------------
def add_donor():
    print(Fore.GREEN + "\n--- Add Donor ---")

    name = input("Name: ").strip()
    contact = input("Contact: ").strip()

    age = int(input("Age: "))
    blood_type = input("Blood Type: ").strip().upper()

    if blood_type not in BLOOD_TYPES:
        animate_message("❌ Invalid blood type.", Fore.RED)
        return

    units = int(input("Units Donated: "))

    donor_id = generate_id("Donors")

    wb = load_workbook_safe()
    ws = wb["Donors"]

    ws.append([
        donor_id, name, contact, age,
        blood_type, units,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ])

    wb.save(DB_FILENAME)

    update_blood_stock(blood_type, units, "add")
    log_history(donor_id, "Donor", name, blood_type, units, "Donated")

    donation_trend.append(units)

    update_excel_dashboard()

    animate_message(f"✅ Donor added! ID: {donor_id}")


def add_patient():
    print(Fore.MAGENTA + "\n--- Add Patient ---")

    name = input("Name: ").strip()
    contact = input("Contact: ").strip()
    age = int(input("Age: "))
    needs_blood = input("Needs Blood? (yes/no): ").lower()

    if needs_blood == "yes":
        blood_type = input("Blood Type: ").strip().upper()

        if blood_type not in BLOOD_TYPES:
            animate_message("❌ Invalid blood type!", Fore.RED)
            return

        units = int(input("Units Needed: "))
    else:
        blood_type = "N/A"
        units = 0

    patient_id = generate_id("Patients")

    wb = load_workbook_safe()
    ws = wb["Patients"]

    ws.append([
        patient_id, name, contact, age,
        blood_type, units,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ])

    wb.save(DB_FILENAME)

    if needs_blood == "yes":
        update_blood_stock(blood_type, units, "subtract")
        log_history(patient_id, "Patient", name, blood_type, units, "Needed")
        usage_trend.append(units)
    else:
        usage_trend.append(0)

    update_excel_dashboard()

    animate_message(f"✅ Patient added! ID: {patient_id}")


def view_sheet(sheet):
    wb = load_workbook_safe()
    ws = wb[sheet]

    headers = [cell.value for cell in ws[1]]
    data = [list(row) for row in ws.iter_rows(min_row=2, values_only=True)]

    print(Fore.CYAN + f"=== {sheet} ===")
    print(Fore.YELLOW + tabulate(data, headers=headers, tablefmt="fancy_grid"))

    input(Fore.CYAN + "Press Enter...")


def search_record():
    sheet = input("Search in (Donors/Patients): ").strip().title()

    if sheet not in ["Donors", "Patients"]:
        animate_message("❌ Invalid sheet.", Fore.RED)
        return

    query = input("Enter Name, ID, or Blood Type: ").lower()

    wb = load_workbook_safe()
    ws = wb[sheet]

    headers = [cell.value for cell in ws[1]]
    results = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(query in str(col).lower() for col in row):
            results.append(list(row))

    if results:
        print(tabulate(results, headers=headers, tablefmt="fancy_grid"))
    else:
        animate_message("❌ No record found.", Fore.RED)

    input("Press Enter...")


def update_record():
    sheet = input("Update in (Donors/Patients): ").strip().title()

    if sheet not in ["Donors", "Patients"]:
        animate_message("❌ Invalid.", Fore.RED)
        return

    record_id = input(f"Enter {sheet[:-1]} ID: ").strip()

    wb = load_workbook_safe()
    ws = wb[sheet]

    for row in ws.iter_rows(min_row=2):
        if str(row[0].value) == record_id:
            row[1].value = input(f"Name [{row[1].value}]: ") or row[1].value
            row[2].value = input(f"Contact [{row[2].value}]: ") or row[2].value
            row[3].value = int(input(f"Age [{row[3].value}]: ") or row[3].value)

            wb.save(DB_FILENAME)
            update_excel_dashboard()

            animate_message("✔ Record updated!")
            break
    else:
        animate_message("❌ Not found.", Fore.RED)

    input("Press Enter...")


def delete_record():
    sheet = input("Delete in (Donors/Patients): ").strip().title()

    if sheet not in ["Donors", "Patients"]:
        animate_message("❌ Invalid.", Fore.RED)
        return

    record_id = input(f"Enter {sheet[:-1]} ID: ").strip()

    wb = load_workbook_safe()
    ws = wb[sheet]

    for row in ws.iter_rows(min_row=2):
        if str(row[0].value) == record_id:
            ws.delete_rows(row[0].row)
            wb.save(DB_FILENAME)
            update_excel_dashboard()

            animate_message("✔ Record deleted!")
            break
    else:
        animate_message("❌ Not found.", Fore.RED)

    input("Press Enter...")

# ---------------------------- Live Dashboard Loop ----------------------------
def live_dashboard():
    try:
        while True:
            display_live_dashboard()

            start = time.time()

            while time.time() - start < DASHBOARD_REFRESH_SECONDS:
                if msvcrt.kbhit():
                    key = msvcrt.getch().decode().lower()

                    if key == "d": add_donor(); break
                    if key == "p": add_patient(); break
                    if key == "h": view_sheet("History"); break
                    if key == "s": search_record(); break
                    if key == "u": update_record(); break
                    if key == "x": delete_record(); break
                    if key == "q": return

    except KeyboardInterrupt:
        return

# ---------------------------- Main ----------------------------
def main():
    initialize_database()

    while True:
        os.system('cls')
        print_banner()

        print(Fore.YELLOW + """
=== Main Menu ===
1. Add Donor
2. Add Patient
3. View Donors
4. View Patients
5. View Blood Stock
6. View History
7. Live Dashboard
8. Exit
""")

        choice = input(Fore.CYAN + "Enter choice: ").strip()

        if choice == "1": add_donor()
        elif choice == "2": add_patient()
        elif choice == "3": view_sheet("Donors")
        elif choice == "4": view_sheet("Patients")
        elif choice == "5": view_sheet("BloodStock")
        elif choice == "6": view_sheet("History")
        elif choice == "7": live_dashboard()
        elif choice == "8":
            animate_message("Goodbye!", Fore.GREEN)
            break
        else:
            animate_message("❌ Invalid choice.", Fore.RED)

        input("Press Enter...")

if __name__ == "__main__":
    main()
