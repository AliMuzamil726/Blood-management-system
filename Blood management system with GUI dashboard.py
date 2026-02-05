import sys
import os
from datetime import datetime, date
from collections import defaultdict

from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QTableWidget, QTableWidgetItem, QLineEdit, QMessageBox,
    QDialog, QFormLayout, QSpinBox, QComboBox, QProgressBar, QGridLayout
)
from PyQt5.QtCore import Qt, QTimer, QPropertyAnimation, QEasingCurve
from PyQt5.QtGui import QFont
import openpyxl
from fpdf import FPDF

# Note: Using backend_qt5agg for PyQt6 requires an environment setup that 
# correctly links Qt5 modules with PyQt6's environment. This often works 
# out-of-the-box but might require specific dependency versions.
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure

# ---------------------------- Configuration ----------------------------
PATIENT_FILE = "patients.xlsx"
DONOR_FILE = "donors.xlsx"
STOCK_FILE = "blood_stock.xlsx"
HISTORY_FILE = "history.xlsx"

# Define low-stock thresholds per blood type
LOW_STOCK_THRESHOLDS = {
    "A+": 5, "A-": 5,
    "B+": 5, "B-": 5,
    "AB+": 3, "AB-": 3,
    "O+": 10, "O-": 10,
}

MAX_STOCK = 100  # Maximum on progress bars for visualization

# ---------------------------- Excel Setup ----------------------------
def init_excel(filename, headers):
    if not os.path.exists(filename):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        wb.save(filename)

# Initialize files
init_excel(PATIENT_FILE, ["ID", "Name", "Age", "Blood Type", "Disease", "Date"])
init_excel(DONOR_FILE, ["ID", "Name", "Age", "Blood Type", "Last Donation Date"])
init_excel(STOCK_FILE, ["Blood Type", "Quantity"])
# HISTORY_FILE has 6 columns
init_excel(HISTORY_FILE, ["DateTime", "Action", "Type", "Name", "BloodType", "Quantity"])

def load_excel(file):
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    # IMPORTANT: The load_excel function returns everything, including the header row.
    return [list(row) for row in ws.iter_rows(values_only=True)]

def save_excel(file, data):
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in data:
        ws.append(row)
    wb.save(file)

def append_excel(file, row):
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    ws.append(row)
    wb.save(file)

# ---------------------------- Animated Stock Bar ----------------------------
class AnimatedStockBar(QWidget):
    def __init__(self, blood_type, quantity, max_quantity=MAX_STOCK):
        super().__init__()
        self.blood_type = blood_type
        self.max_quantity = max_quantity

        self.layout = QVBoxLayout()
        self.label = QLabel(f"{blood_type}: {quantity} units")
        self.label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.progress = QProgressBar()
        self.progress.setMaximum(self.max_quantity)
        self.progress.setValue(quantity)
        self.progress.setStyleSheet(self.get_style(quantity))
        self.layout.addWidget(self.label)
        self.layout.addWidget(self.progress)
        self.setLayout(self.layout)

    def get_style(self, value):
        # Flash red if below threshold
        threshold = LOW_STOCK_THRESHOLDS.get(self.blood_type, 0)
        if value <= threshold:
            return """
                QProgressBar {
                    border: 2px solid grey;
                    border-radius: 5px;
                    text-align: center;
                }
                QProgressBar::chunk {
                    background-color: red;
                    width: 20px;
                }
            """
        else:
            return """
                QProgressBar {
                    border: 2px solid grey;
                    border-radius: 5px;
                    text-align: center;
                }
                QProgressBar::chunk {
                    background: qlineargradient(
                        x1:0, y1:0, x2:1, y2:0,
                        stop:0 green, stop:1 blue
                    );
                }
            """

    def animate_to(self, new_value):
        anim = QPropertyAnimation(self.progress, b"value", self)
        anim.setDuration(800)
        anim.setStartValue(self.progress.value())
        anim.setEndValue(new_value)
        anim.setEasingCurve(QEasingCurve.Type.InOutQuad)
        anim.start()
        self.progress.setStyleSheet(self.get_style(new_value))
        self.label.setText(f"{self.blood_type}: {new_value} units")
        # Keep a reference to prevent garbage collection
        self._anim = anim

# ---------------------------- Trend Graph Canvas ----------------------------
class TrendCanvas(FigureCanvas):
    def __init__(self, parent=None):
        fig = Figure(figsize=(5, 3))
        self.axes = fig.add_subplot(111)
        super().__init__(fig)
        self.setParent(parent)
        self.axes.set_title("Donations vs Usage (Last 30 days)")
        self.axes.set_xlabel("Date")
        self.axes.set_ylabel("Units")
        fig.tight_layout()

    def update_plot(self, history_data):
        # history_data: list of rows from HISTORY_FILE
        daily = defaultdict(lambda: {"donated": 0, "used": 0})
        
        # Start iteration from the second row to skip the header
        for row in history_data[1:]:
            # Check for exactly 6 elements before unpacking (robustness)
            if len(row) != 6:
                print(f"Skipping malformed history row with length {len(row)}: {row}")
                continue
                
            # Unpacking 6 elements from a row with 6 columns
            dt_str, action, typ, name, btype, qty = row
            
            try:
                dt = datetime.fromisoformat(dt_str)
                d = dt.date()
                quantity = int(qty)
            except Exception:
                # Skip rows with malformed dates or quantities
                continue
            
            if action == "Add Donor":
                daily[d]["donated"] += quantity
            elif action == "Add Patient":
                daily[d]["used"] += quantity

        # Sort days
        days = sorted(daily.keys())
        donated = [daily[d]["donated"] for d in days]
        used = [daily[d]["used"] for d in days]

        self.axes.clear()
        if days:
            self.axes.plot(days, donated, label="Donated", color="green", marker="o")
            self.axes.plot(days, used, label="Used", color="red", marker="x")
            self.axes.legend()
            self.axes.set_xlabel("Date")
            self.axes.set_ylabel("Units")
            self.axes.set_title("Donations vs Usage (Last ~30 days)")
        else:
            self.axes.text(0.5, 0.5, "No data", horizontalalignment='center', verticalalignment='center')
        self.draw()

# ---------------------------- Main Application ----------------------------
class BloodManagementSystem(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("ICU Blood Management System (PyQt6)")
        self.setGeometry(50, 50, 1400, 800)
        self.layout = QVBoxLayout()
        self.setLayout(self.layout)

        # Title
        title = QLabel("🩸 ICU Blood Management System 🏥")
        title.setFont(QFont("Arial", 20, QFont.Weight.Bold))
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.layout.addWidget(title)

        # Buttons
        btn_layout = QHBoxLayout()
        self.layout.addLayout(btn_layout)
        buttons = [
            ("Add Patient", self.add_patient),
            ("View Patients", self.view_patients),
            ("Add Donor", self.add_donor),
            ("View Donors", self.view_donors),
            ("View History", self.view_history),
            ("Export PDF Report", self.export_pdf)
        ]
        for text, func in buttons:
            b = QPushButton(text)
            b.clicked.connect(func)
            b.setStyleSheet("background-color:#66ccff; font-weight:bold; font-size:14px;")
            btn_layout.addWidget(b)

        # Search bar
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search by Name / Blood Type...")
        self.search_input.textChanged.connect(self.update_search)
        self.layout.addWidget(self.search_input)

        # Main Table
        self.table = QTableWidget()
        self.layout.addWidget(self.table)

        # Dashboard layout (stock bars + metrics + trend graph)
        self.dashboard_layout = QVBoxLayout()
        self.layout.addLayout(self.dashboard_layout)

        # Stock bars
        self.stock_grid = QGridLayout()
        self.dashboard_layout.addLayout(self.stock_grid)
        self.stock_bars = {}
        blood_types = ["A+", "A-", "B+", "B-", "AB+", "AB-", "O+", "O-"]
        for i, bt in enumerate(blood_types):
            bar = AnimatedStockBar(bt, 0)
            self.stock_bars[bt] = bar
            row = i // 4
            col = i % 4
            self.stock_grid.addWidget(bar, row, col)

        # Metrics panel (labels)
        self.metrics_layout = QHBoxLayout()
        self.dashboard_layout.addLayout(self.metrics_layout)
        self.patient_counter = QLabel("Total Patients: 0")
        self.donor_counter = QLabel("Total Donors: 0")
        self.daily_don_label = QLabel("Today’s Donations: 0")
        self.daily_use_label = QLabel("Today’s Usage: 0")
        for lbl in [self.patient_counter, self.donor_counter, self.daily_don_label, self.daily_use_label]:
            lbl.setFont(QFont("Arial", 14, QFont.Weight.Bold))
            lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            self.metrics_layout.addWidget(lbl)

        # Alerts summary
        self.alerts_label = QLabel("")
        self.alerts_label.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        self.alerts_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.dashboard_layout.addWidget(self.alerts_label)

        # Trend Graph
        self.trend_canvas = TrendCanvas(self)
        self.dashboard_layout.addWidget(self.trend_canvas)

        # Refresh Timer
        self.timer = QTimer()
        self.timer.timeout.connect(self.refresh_dashboard)
        self.timer.start(2000)  # refresh every 2 seconds

        self.current_view = None
        self.refresh_dashboard()

    # ---------------------------- CRUD: Add Patient / Donor ----------------------------
    def add_patient(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Add Patient")
        layout = QFormLayout(dialog)

        name_input = QLineEdit()
        age_input = QSpinBox()
        age_input.setRange(0, 120)
        blood_input = QComboBox()
        blood_input.addItems(list(self.stock_bars.keys()))
        disease_input = QLineEdit()

        layout.addRow("Name:", name_input)
        layout.addRow("Age:", age_input)
        layout.addRow("Blood Type:", blood_input)
        layout.addRow("Disease:", disease_input)

        btn_add = QPushButton("Add")
        layout.addRow(btn_add)

        def save_patient():
            name = name_input.text().strip()
            if not name:
                QMessageBox.warning(dialog, "Error", "Name cannot be empty.")
                return
            bt = blood_input.currentText()
            stock = load_excel(STOCK_FILE)
            # reduce stock by 1 (usage)
            for row in stock[1:]:
                if row[0] == bt:
                    if row[1] <= 0:
                        QMessageBox.warning(dialog, "Error", f"No stock available for {bt}.")
                        return
                    row[1] -= 1
                    break
            else:
                QMessageBox.warning(dialog, "Error", f"No stock record for {bt}.")
                return

            save_excel(STOCK_FILE, stock)

            pdata = load_excel(PATIENT_FILE)
            new_id = len(pdata)
            row = [new_id, name, age_input.value(), bt, disease_input.text(), datetime.now().strftime('%Y-%m-%d')] # Changed to YYYY-MM-DD for simpler viewing
            append_excel(PATIENT_FILE, row)
            # Append 6 values to HISTORY_FILE
            append_excel(HISTORY_FILE, [datetime.now().isoformat(), "Add Patient", "Patient", name, bt, 1])

            QMessageBox.information(dialog, "Success", "Patient Added (Blood usage recorded).")
            dialog.close()
            self.refresh_dashboard()

        btn_add.clicked.connect(save_patient)
        dialog.exec()

    def add_donor(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Add Donor")
        layout = QFormLayout(dialog)

        name_input = QLineEdit()
        age_input = QSpinBox()
        age_input.setRange(0, 120)
        blood_input = QComboBox()
        blood_input.addItems(list(self.stock_bars.keys()))
        last_donation_input = QLineEdit()
        last_donation_input.setPlaceholderText("YYYY-MM-DD")

        layout.addRow("Name:", name_input)
        layout.addRow("Age:", age_input)
        layout.addRow("Blood Type:", blood_input)
        layout.addRow("Last Donation:", last_donation_input)

        btn_add = QPushButton("Add")
        layout.addRow(btn_add)

        def save_donor():
            name = name_input.text().strip()
            if not name:
                QMessageBox.warning(dialog, "Error", "Name cannot be empty.")
                return
            bt = blood_input.currentText()
            # update stock by +1 (donation)
            stock = load_excel(STOCK_FILE)
            updated = False
            for row in stock[1:]:
                if row[0] == bt:
                    row[1] += 1
                    updated = True
                    break
            if not updated:
                stock.append([bt, 1])
            save_excel(STOCK_FILE, stock)

            ddata = load_excel(DONOR_FILE)
            new_id = len(ddata)
            row = [new_id, name, age_input.value(), bt, last_donation_input.text()]
            append_excel(DONOR_FILE, row)
            # Append 6 values to HISTORY_FILE
            append_excel(HISTORY_FILE, [datetime.now().isoformat(), "Add Donor", "Donor", name, bt, 1])

            QMessageBox.information(dialog, "Success", "Donor Added (Stock updated).")
            dialog.close()
            self.refresh_dashboard()

        btn_add.clicked.connect(save_donor)
        dialog.exec()

    # ---------------------------- View (Patients, Donors, History) ----------------------------
    def view_patients(self):
        self.current_view = "patient"
        self.populate_table(load_excel(PATIENT_FILE))

    def view_donors(self):
        self.current_view = "donor"
        self.populate_table(load_excel(DONOR_FILE))

    def view_history(self):
        self.current_view = "history"
        self.populate_table(load_excel(HISTORY_FILE))

    # ---------------------------- Table population & Search ----------------------------
    def populate_table(self, data):
        self.table.clear()
        if not data:
            self.table.setRowCount(0)
            self.table.setColumnCount(0)
            return
        
        headers = data[0]
        rows = data[1:]

        self.table.setRowCount(len(rows))
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels([str(h) for h in headers])

        for i, row in enumerate(rows):
            for j, val in enumerate(row):
                item = QTableWidgetItem(str(val))
                item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)
                self.table.setItem(i, j, item)
        self.table.resizeColumnsToContents()


    def update_search(self):
        if not self.current_view:
            return
        query = self.search_input.text().lower()
        data = []
        if self.current_view == "patient":
            data = load_excel(PATIENT_FILE)
        elif self.current_view == "donor":
            data = load_excel(DONOR_FILE)
        elif self.current_view == "history":
            data = load_excel(HISTORY_FILE)
            
        if not data:
             self.populate_table([])
             return
             
        # Separate headers and data rows for filtering
        headers = data[0]
        data_rows = data[1:]
        
        filtered_rows = [row for row in data_rows if any(query in str(cell).lower() for cell in row)]
        
        # Repack the filtered data with headers for population
        self.populate_table([headers] + filtered_rows)

    # ---------------------------- Dashboard Refresh ----------------------------
    def refresh_dashboard(self):
        # Refresh stock bars
        stock_data = load_excel(STOCK_FILE)
        if len(stock_data) == 1:
            # only header, initialize
            for bt in self.stock_bars:
                append_excel(STOCK_FILE, [bt, 0])
            stock_data = load_excel(STOCK_FILE)

        for row in stock_data[1:]:
            bt, qty = row
            if bt in self.stock_bars:
                self.stock_bars[bt].animate_to(qty)

        # Update counters
        # Subtract 1 for the header row
        total_patients = len(load_excel(PATIENT_FILE)) - 1
        total_donors = len(load_excel(DONOR_FILE)) - 1
        self.patient_counter.setText(f"Total Patients: {total_patients}")
        self.donor_counter.setText(f"Total Donors: {total_donors}")

        # Daily donations / usage
        today = date.today()
        hist = load_excel(HISTORY_FILE)
        donated = 0
        used = 0
        for row in hist[1:]:
            # Check for exactly 6 elements before unpacking (robustness)
            if len(row) != 6:
                continue
                
            # FIX: The original code used 4 placeholders, but history has 6 columns
            # The correct unpacking for all 6 columns, ignoring the 3rd, 4th, 5th, is:
            dt_str, action, typ, name, btype, qty = row
            
            try:
                dt = datetime.fromisoformat(dt_str)
                quantity = int(qty)
            except:
                continue
            
            if dt.date() == today:
                if action == "Add Donor":
                    donated += quantity
                elif action == "Add Patient":
                    used += quantity
        
        self.daily_don_label.setText(f"Today’s Donations: {donated}")
        self.daily_use_label.setText(f"Today’s Usage: {used}")

        # Alerts summary
        alerts = []
        for bt, bar in self.stock_bars.items():
            qty = bar.progress.value()
            if qty <= LOW_STOCK_THRESHOLDS.get(bt, 0):
                alerts.append(f"{bt} low ({qty})")
        if alerts:
            self.alerts_label.setText("⚠ Low stock: " + ", ".join(alerts))
            self.alerts_label.setStyleSheet("color: red;")
        else:
            self.alerts_label.setText("All blood types stock is healthy.")
            self.alerts_label.setStyleSheet("color: green;")

        # Update trend graph
        self.trend_canvas.update_plot(hist)

    # ---------------------------- PDF Export ----------------------------
    def export_pdf(self):
        if not self.current_view:
            QMessageBox.warning(self, "Error", "Please view Patients, Donors or History first.")
            return
        data = []
        title = ""
        if self.current_view == "patient":
            data = load_excel(PATIENT_FILE)
            title = "Patient Report"
        elif self.current_view == "donor":
            data = load_excel(DONOR_FILE)
            title = "Donor Report"
        elif self.current_view == "history":
            data = load_excel(HISTORY_FILE)
            title = "History Report"

        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()
        pdf.set_font("Arial", "B", 16)
        pdf.cell(0, 10, txt=title, ln=True, align="C")
        pdf.ln(5)
        pdf.set_font("Arial", size=10) # Reduced font size for better fit

        # Print data, excluding the header row
        if data:
            # Table headers
            pdf.set_fill_color(200, 220, 255)
            pdf.set_font("Arial", "B", 10)
            
            col_widths = [20, 40, 15, 20, 30, 45] # Adjusted widths for 6 columns

            headers = [str(h) for h in data[0]]
            # Determine appropriate column width (simple uniform example)
            width = 190 / len(headers) 
            
            # Print Headers
            for i, header in enumerate(headers):
                pdf.cell(width, 7, header, 1, 0, 'C', 1)
            pdf.ln()

            # Table rows
            pdf.set_font("Arial", size=10)
            for row in data[1:]:
                for i, val in enumerate(row):
                    pdf.cell(width, 6, str(val), 1, 0, 'L')
                pdf.ln()

        # Add trend summary
        pdf.add_page()
        pdf.set_font("Arial", "B", 14)
        pdf.cell(0, 10, txt="Blood Stock & Usage Summary", ln=True)
        pdf.ln(3)
        pdf.set_font("Arial", size=12)
        stock = load_excel(STOCK_FILE)
        for row in stock[1:]:
            bt, qty = row
            pdf.cell(0, 8, txt=f"{bt}: {qty} units", ln=True)

        filename = f"{title.replace(' ', '_').replace('/', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        pdf.output(filename)
        QMessageBox.information(self, "Success", f"Report saved to {filename}")

# ---------------------------- Run Application ----------------------------
if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = BloodManagementSystem()
    window.show()
    sys.exit(app.exec())